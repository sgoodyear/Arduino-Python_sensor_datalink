import datetime
import time
import serial
import sys
import select
import os
import openpyxl as xl


def B2I(in_bytes):                                            # Reconstructs 4 bytes of data into a long int
    value = (in_bytes[0] << 24) | (in_bytes[1] << 16) | (in_bytes[2] << 8) | in_bytes[3]
    return value


def InputTimeout(prompt, timeout):
    prompt_time = time.time()
    while time.time() - prompt_time < timeout:
        print(prompt, end='', flush=True)
        if sys.stdin in select.select([sys.stdin], [], [], timeout)[0]:
            return sys.stdin.readline().strip()
    return None


# Once the sensors have been installed and wired into the multiplexer arrays, coordinate locations will be assigned here
# Sensor layout and coordinate system are defined here: https://docs.google.com/presentation/d/1iU9-4sDkxN5r5blLVAVAElmQhg7pdddcT2UxLzKhD1Q/edit?usp=sharing
Airspeed_Sensor_Coordinates = ["00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                               "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00"]
Gas_Sensor_Coordinates = ["00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                          "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                          "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                          "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                          "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                          "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                          "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                          "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                          "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                          "00-00", "00-00", "00-00"]
Temperature_Sensor_Coordinates = ["00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                                  "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                                  "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                                  "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                                  "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                                  "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                                  "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                                  "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                                  "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00",
                                  "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00", "00-00"]

now = datetime.datetime.now()                                  # Setting up info for timestamping
filename = now.strftime("%Y-%m-%d_%H-%M") + "_data.xlsx"       # Setting up unique filename for exported data

mainWB = xl.Workbook(write_only=True)                          # Create spreadsheet to store data, using write-only mode in case the datasets get big
datasheet = mainWB.create_sheet("Data")                        # Name the sheet we want to work on
# del mainWB["sheet"]                                          # Delete artefact (only need it when not using write_only mode)

n = 0                                                          # Start master index at 0

ser = serial.Serial('/dev/ttyUSB0', 1000000)                   # Open serial port
buffer, data = b"", b""
BadData = False                                                # Initializing variables during startup
Header = False
cleanAData, cleanTData, cleanSData = [], [], []
AirFlowData, refinedAData, refinedTData, refinedSData = [], [], [], []

try:  # Main code runs here
    print("\033[94m" + "Serial connection opened" + "\033[0m")
    while True:
        intro = ser.readline().decode().strip()                # For the 'mirroring' before actual data is coming in, use built-in decoding
        if intro == "Nano Pass-Through ready":
            print(intro)
            break
        else:
            print(intro)
    del intro                                                                             # Remove the string from memory to keep space free
    x = 0
    time.sleep(1)

    print()
    print()
    print("Mode " + "\033[32m" + "1" + "\033[0m" + ": Save at End (default)")
    print("Mode " + "\033[32m" + "2" + "\033[0m" + ": Save Iteratively")
    SetMode = InputTimeout("Enter the number of your selection: ", 10)      # Should the program save every time or just once at the end?
    if SetMode is not None:
        if SetMode == "1":
            print("Running in default mode.")
        elif SetMode == "2":
            print("Running in iterative save mode. (You may see increased memory usage)")
        else:
            print("Mode selection not recognized, setting to default")
    else:
        print("Timed out. Running in default mode.")
        SetMode = 1
    print()
    print()

    start_time = time.time()                                                              # Record starting time for data indexing
    ser.write(b"\x47")                                                                    # Request first data set
    current_time = time.time() - start_time
    print("Startup time: " + str(current_time))

    last_time = current_time

    while True:
        data = ser.read()                                                                 # As data comes in, do not apply decoding since everything is meant to be in binary
        buffer += data                                                                    # Put it into a buffer for short-term storage

        if buffer.endswith(b"?????"):                                                     # Check if the Arduino returned an error
            print("\033[93m" + "Warning: Nano pass-thru returned an error" + "\033[0m")
            buffer = b""                                                                  # If so, make sure to reset the buffer
            BadData = True
            time.sleep(1)

        if buffer.endswith(b"~~~~~"):                                                     # And if the terminating symbol comes in...
            SplitData = list(buffer)                                                      # Split up the byte array into list items that are easier to work with
            if SplitData[0] == 65 and len(SplitData) == 120:                              # Check if it starts with 0x65 (DEC for "A" for airspeed) and make sure all of it is there
                cleanAData = [item for index, item in enumerate(SplitData) if (index + 1) % 5 != 1]  # Clean out the sensor type identifiers
                cleanAData = cleanAData[:-4]                                              # Remove the last remaining bit of the 5-character terminator symbol
                for i in range(0, len(cleanAData), 4):
                    convertedChunk = B2I(cleanAData[i:i+4])                               # Convert the data from binary to integers, in blocks of 4 bytes
                    refinedAData.append(convertedChunk / 10000)                           # Convert integer to float using the opposite operation as the Arduino Mega made
                print(refinedAData)
                AirFlowData = [airspeed * 0.00112903 for airspeed in refinedAData]        # Spin off a new list for airflow by multiplying by cross-sectional area

            elif SplitData[0] == 83 and len(SplitData) == 470:                            # Check for DEC "S" for smoke data and make sure all of it is there
                cleanSData = [item for index, item in enumerate(SplitData) if (index + 1) % 5 != 1]  # Clean out the sensor type identifiers
                cleanSData = cleanSData[:-4]                                              # Remove the last of the 5-character terminator symbol
                for i in range(0, len(cleanSData), 4):
                    convertedChunk = B2I(cleanSData[i:i+4])                               # Convert the data from binary to integers, in blocks of 4 bytes
                    refinedSData.append(convertedChunk)                                   # No need to convert to float here, it is intended to be an integer
                print(refinedSData)

            elif SplitData[0] == 84 and len(SplitData) == 450:                            # Check for DEC "T" for temperature data and make sure all of it is there
                cleanTData = [item for index, item in enumerate(SplitData) if (index + 1) % 5 != 1]  # Clean out the sensor type identifiers
                cleanTData = cleanTData[:-4]                                              # Remove the last of the 5-character terminator symbol
                for i in range(0, len(cleanTData), 4):
                    convertedChunk = B2I(cleanTData[i:i+4])                               # Convert the data from binary to integers, in blocks of 4 bytes
                    refinedTData.append(convertedChunk / 10000)                           # Convert integer to float using the opposite operation as the Arduino Mega made
                print(refinedTData)

            else:
                print("\033[93m" + "Warning: Bad serial TX/RX. Suspected bad data set will be dumped." + "\033[0m")
                BadData = True                                                            # If something isn't quite right about the data, don't let it get saved to the spreadsheet later

            buffer = b""                                                                  # Purge buffer for next set of incoming data
            x += 1
            if x == 3:                                                                    # And after all 3 sets of data have arrived...
                x = 0                                                                     # Reset the count
                current_time = time.time() - start_time                                   # Check the time
                print("Timestamp: " + str(int(current_time)) + " seconds")                # Print timestamp for user
                print()

                if not Header:                                                            # If the spreadsheet was not already given a header...
                    HeaderList = [str(filename)] + ["Airflow (m^3/s)" for _ in AirFlowData] + ["Air Velocity (m/s)" for _ in refinedAData] + ["Temperature (C)" for _ in refinedTData] + ["Gas (ppm)" for _ in refinedSData]
                    datasheet.append(HeaderList)                                          # Create one based on the amount of data being received
                    HeaderList = ["Time (seconds)"] + Airspeed_Sensor_Coordinates + Airspeed_Sensor_Coordinates + Temperature_Sensor_Coordinates + Gas_Sensor_Coordinates
                    datasheet.append(HeaderList)
                    del HeaderList                                                        # Clean up memory
                    Header = True                                                         # Now it has a header

                if not BadData:                                                           # If the data was acceptable, add it to the spreadsheet
                    ExportList = [int(current_time)] + AirFlowData + refinedAData + refinedTData + refinedSData
                    datasheet.append(ExportList)
                elif BadData:
                    print("\033[93m" + "Suspected bad data set was dumped." + "\033[0m")
                BadData = False                                                           # Reset data integrity flag for next set of incoming data

                refinedAData, refinedTData, refinedSData = [], [], []                     # Make room for new data decoding
                cleanAData, cleanTData, cleanSData = [], [], []

                if SetMode == "2":
                    mainWB.save(filename)                                                 # Save the spreadsheet
                    mainWB.close()
                    mainWB = xl.load_workbook(filename=filename)                          # Re-load the spreadsheet
                    datasheet = mainWB["Data"]
                    print("Spreadsheet updated.")

                while current_time - last_time < 10:                                      # Regardless of how long it takes data to come in, make sure there is a constant timing
                    time.sleep(0.1)
                    current_time = time.time() - start_time
                    print('.', end='')

                last_time = current_time
                print()
                print("Requesting new data...")
                ser.write(b"\x47")                                                        # Send the "request data" command to the Nano

        current_time = time.time() - start_time                                           # Update clock
        n += 1                                                                            # At the end of the loop, advance the master index


except KeyboardInterrupt:                                                                 # Gracefully save data and shut down the program instead of crashing
    print("\033[93m" + "[Keyboard interrupt]" + "\033[0m")
    if ser.is_open:
        ser.close()                                                                       # Close the serial connection
        print("\033[94m" + "Serial connection closed" + "\033[0m")

    try:
        print("Saving spreadsheet...")
        mainWB.save(filename)                                                             # Save the spreadsheet
        mainWB.close()
        print("Spreadsheet saved: " + "\033[32m" + "/home/simrigcontrol/PycharmProjects/Arduino-Python_sensor_datalink/" + filename + "\033[0m")
    except FileNotFoundError:                                                             # ... Unless it was already saved
        print("Spreadsheet was already saved: " + "\033[32m" + "/home/simrigcontrol/PycharmProjects/Arduino-Python_sensor_datalink/" + filename + "\033[0m")

    final = InputTimeout("Print 'del' to delete the spreadsheet. Do nothing for 10 seconds or enter anything else to keep the spreadsheet:", 10)
    if final == "del":
        os.remove("/home/simrigcontrol/PycharmProjects/Arduino-Python_sensor_datalink/" + filename)
        print("File deleted.")
        time.sleep(2.5)
    else:
        print("Exiting program.")
        time.sleep(2.5)
    sys.exit(0)                                                                           # Show "success" exit code
